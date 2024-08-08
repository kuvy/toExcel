import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.worksheet.page import PageMargins

def read_file(file_path):
    data = []
    with open(file_path, 'r', encoding='windows-1251') as file:
        for line in file:
            parts = line.strip().split()
            if len(parts) == 5:
                last_name, first_name, patronymic, liters, date_str = parts
                try:
                    # Проверяем правильность формата даты
                    date = datetime.datetime.strptime(date_str, '%d.%m.%Y')
                    # Проверяем правильность формата числа литров
                    liters = float(liters)
                    data.append((last_name, first_name, patronymic, liters, date_str))
                except (ValueError, IndexError):
                    print(f"Ошибка в строке: {line.strip()}")
            else:
                print(f"Неверный формат строки: {line.strip()}")
    return data

# Функция для очистки существующего листа таблицы
def create_or_clear_sheet(workbook, sheet_name):
    if sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        for row in sheet.iter_rows():
            for cell in row:
                cell.value = None
    else:
        sheet = workbook.create_sheet(sheet_name)
    return sheet

def insert_into_excel(data, excel_path, sheet_name='Sheet1'):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_name

    # Настройка шапки таблицы
    header = ["ФИО", "Литры"]
    for col_num, header_title in enumerate(header, 1):
        cell = sheet.cell(row=1, column=col_num)
        cell.value = header_title
        cell.font = Font(name='Arial', bold=True, size=12)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Начинаем вставку с первой пустой строки
    start_row = sheet.max_row + 1
    for row_num, entry in enumerate(data, start=start_row):
        last_name, first_name, patronymic, liters, date_str = entry

        sheet[f'A{row_num}'] = f"{last_name} {first_name} {patronymic}"
        sheet[f'B{row_num}'] = liters

        # Пример форматирования
        for col in 'AB':
            cell = sheet[f'{col}{row_num}']
            cell.font = Font(name='Arial', size=12)
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # Настройка параметров печати
    sheet.print_title_rows = '1:1'  # Повтор первой строки на каждой странице
    sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE
    sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
    sheet.page_margins = PageMargins(left=0.5, right=0.5, top=0.75, bottom=0.75, header=0.3, footer=0.3)

    workbook.save(excel_path)
    print(f"Данные успешно вставлены в {excel_path}")

if __name__ == '__main__':
    file_path = 'UserReport.txt'
    excel_file_path = 'Result.xlsx'
    result = read_file(file_path)
    for entry in result:
        print(entry)
    insert_into_excel(result, excel_file_path)

